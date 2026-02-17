#!/usr/bin/env python3
"""
Dormant Account Analysis
Produces an Excel workbook with:
  Sheet 1 - Dormant Definition: return rate by inactivity threshold
  Sheet 2 - Segment Analysis: return rates by segment/sub-segment
  Sheet 3 - Monthly Model: active/dormant/reactivated/new with variable threshold
  Sheet 4 - Reference Data: underlying data powering the dropdown

Assumptions:
  - Rows with $0 invoiced AND 0 qty are excluded (no revenue = not a purchase)
  - "Dormant" means no qualifying purchase within the lookback window
  - Segments and sub-segments are dynamic (discovered from the data)

Usage:
  python dormant_analysis.py <account_csv> <transaction_csv> <output_xlsx> [--start-year YYYY]
"""

import argparse
import sys
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FONT_WHITE = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
DATA_FONT = Font(name='Arial', size=10)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='2F5496')
SUBTITLE_FONT = Font(name='Arial', bold=True, size=12, color='2F5496')
PCT_FMT = '0.0%'
NUM_FMT = '#,##0'
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
HIGHLIGHT_FILL = PatternFill('solid', fgColor='FFFF00')
GREEN_CELL = PatternFill('solid', fgColor='C6EFCE')
YELLOW_CELL = PatternFill('solid', fgColor='FFEB9C')
RED_CELL = PatternFill('solid', fgColor='FFC7CE')
ALT_ROW = PatternFill('solid', fgColor='F2F2F2')
HILITE_ROW = PatternFill('solid', fgColor='E2EFDA')

THRESHOLDS = [6, 9, 12, 15, 18, 24]
RETURN_THRESHOLDS = [3, 6, 9, 12, 15, 18, 21, 24, 30, 36]
GAP_CUM_MONTHS = [1, 2, 3, 6, 9, 12, 18, 24, 36]


def load_and_clean(acct_path, txn_path):
    acct = pd.read_csv(acct_path, dtype=str)
    txn = pd.read_csv(txn_path, dtype=str)

    # Remove grand-total rows
    if 'IsGrandTotalRowTotal' in acct.columns:
        acct = acct[acct['IsGrandTotalRowTotal'] == 'False'].copy()
    if 'IsGrandTotalRowTotal' in txn.columns:
        txn = txn[txn['IsGrandTotalRowTotal'] == 'False'].copy()

    txn['SumTY_Invoiced_Amount'] = pd.to_numeric(txn['SumTY_Invoiced_Amount'], errors='coerce')
    txn['TY_Net_Qty'] = pd.to_numeric(txn['TY_Net_Qty'], errors='coerce')
    txn['Year'] = txn['Year'].astype(int)
    txn['MonthNo'] = txn['MonthNo'].astype(int)
    txn['Day'] = txn['Day'].astype(int)

    # Exclude $0 invoiced AND 0 qty (no revenue)
    txn = txn[~((txn['SumTY_Invoiced_Amount'] == 0) & (txn['TY_Net_Qty'] == 0))].copy()

    txn['order_date'] = pd.to_datetime(
        txn[['Year', 'MonthNo', 'Day']].rename(columns={'Year': 'year', 'MonthNo': 'month', 'Day': 'day'})
    )
    txn['order_month'] = txn['order_date'].dt.to_period('M')

    return acct, txn


def analysis1_return_rates(txn, today):
    """Gap analysis: what % of accounts return after X months of inactivity."""
    acct_months = txn.groupby('Customer Number')['order_month'].apply(
        lambda x: sorted(x.unique())
    ).to_dict()

    all_gaps = []
    last_gaps = []

    for cust, months in acct_months.items():
        for i in range(1, len(months)):
            gap = (months[i] - months[i - 1]).n
            if gap > 0:
                all_gaps.append(gap)
        months_since = (today - months[-1]).n
        if months_since > 0:
            last_gaps.append(months_since)

    results = []
    for threshold in RETURN_THRESHOLDS:
        returned = sum(1 for g in all_gaps if g >= threshold)
        not_returned = sum(1 for g in last_gaps if g >= threshold)
        total = returned + not_returned
        results.append({
            'threshold': threshold,
            'total': total,
            'returned': returned,
            'not_returned': not_returned,
        })

    # Cumulative gap distribution
    all_combined = all_gaps + last_gaps
    gap_series = pd.Series(all_combined)
    total_gaps = len(gap_series)
    cum = gap_series.value_counts().sort_index().cumsum()
    gap_cum = []
    for m in GAP_CUM_MONTHS:
        if m in cum.index:
            gap_cum.append((m, cum[m] / total_gaps))
        else:
            nearest = cum.index[cum.index <= m]
            if len(nearest) > 0:
                gap_cum.append((m, cum[nearest[-1]] / total_gaps))

    return results, gap_cum


def analysis1_segment_rates(txn, today):
    """Return rates broken out by Segment and Sub Segment."""
    acct_seg = txn.sort_values('order_date').groupby('Customer Number').agg(
        last_segment=('Segment', 'last'),
        last_sub_segment=('Sub Segment', 'last'),
    ).reset_index()

    acct_months = txn.groupby('Customer Number')['order_month'].apply(
        lambda x: sorted(x.unique())
    ).reset_index()
    acct_months.columns = ['Customer Number', 'purchase_months']
    acct_months = acct_months.merge(acct_seg, on='Customer Number', how='left')

    def build_gaps(group_col):
        records = []
        for _, row in acct_months.iterrows():
            cust = row['Customer Number']
            grp = row[group_col]
            months = row['purchase_months']
            for i in range(1, len(months)):
                gap = (months[i] - months[i - 1]).n
                if gap > 0:
                    records.append({'group': grp, 'gap': gap, 'returned': True})
            ms = (today - months[-1]).n
            if ms > 0:
                records.append({'group': grp, 'gap': ms, 'returned': False})
        return pd.DataFrame(records)

    seg_thresholds = [6, 9, 12, 15, 18, 24]

    def compute_rates(df):
        groups = df.groupby('group').size().sort_values(ascending=False)
        top_groups = groups[groups >= 20].index.tolist()
        rows = []
        for grp in top_groups:
            gdata = df[df['group'] == grp]
            total = len(gdata[gdata['gap'] >= 1])
            rates = {}
            for t in seg_thresholds:
                at_risk = gdata[gdata['gap'] >= t]
                if len(at_risk) > 0:
                    rates[t] = at_risk['returned'].mean()
                else:
                    rates[t] = None
            rows.append({'group': grp, 'total': total, **{f'{t}mo': rates[t] for t in seg_thresholds}})
        return rows

    seg_gaps = build_gaps('last_segment')
    sub_gaps = build_gaps('last_sub_segment')

    return compute_rates(seg_gaps), compute_rates(sub_gaps)


def analysis2_monthly_model(txn, start_year):
    """Monthly active/dormant/reactivated/new model for multiple thresholds."""
    acct_active_months = txn.groupby('Customer Number')['order_month'].apply(set).to_dict()
    acct_first = txn.groupby('Customer Number')['order_month'].min().to_dict()
    all_customers = list(acct_first.keys())

    max_month = txn['order_month'].max()
    months_range = pd.period_range(f'{start_year}-01', str(max_month), freq='M')

    results = {}
    for threshold in THRESHOLDS:
        prev_active_set = set()
        prev_dormant_set = set()
        monthly_data = []

        for month in months_range:
            window_start = month - threshold
            active_set = set()
            for c in all_customers:
                if acct_first[c] > month:
                    continue
                for am in acct_active_months.get(c, set()):
                    if window_start < am <= month:
                        active_set.add(c)
                        break

            ever_purchased = {c for c in all_customers if acct_first[c] <= month}
            dormant_set = ever_purchased - active_set
            new_set = {c for c in all_customers if acct_first[c] == month}
            newly_dormant_set = prev_active_set - active_set
            reactivated_set = (prev_dormant_set & active_set) - new_set

            monthly_data.append({
                'month': str(month),
                'active': len(active_set),
                'dormant_total': len(dormant_set),
                'newly_dormant': len(newly_dormant_set),
                'reactivated': len(reactivated_set),
                'new_accounts': len(new_set),
                'ever_purchased': len(ever_purchased),
            })

            prev_active_set = active_set
            prev_dormant_set = dormant_set

        results[threshold] = monthly_data
        print(f"  Threshold {threshold}mo computed ({len(monthly_data)} months)")

    return results, list(months_range)


# ── Excel Builder ───────────────────────────────────────────────────────────

def build_sheet1(wb, return_results, gap_cum):
    ws = wb.active
    ws.title = "Dormant Definition"

    ws['A1'] = "Analysis 1: When Is an Account Truly Dormant?"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')

    ws['A3'] = ("This analysis examines purchase gaps across all accounts to determine "
                "the optimal inactivity threshold for classifying accounts as dormant. "
                "Excludes $0 invoiced / 0 qty rows (no revenue = not a purchase).")
    ws['A3'].font = Font(name='Arial', size=10, italic=True)
    ws.merge_cells('A3:G3')

    ws['A5'] = "Return Rate by Months of Inactivity"
    ws['A5'].font = SUBTITLE_FONT

    headers = ['Months Inactive', 'Accounts at Risk', 'Returned', 'Did Not Return',
               '% Returned', '% Did Not Return']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for i, r in enumerate(return_results):
        row = 7 + i
        ws.cell(row=row, column=1, value=r['threshold']).font = DATA_FONT
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=r['total']).font = DATA_FONT
        ws.cell(row=row, column=2).number_format = NUM_FMT
        ws.cell(row=row, column=3, value=r['returned']).font = DATA_FONT
        ws.cell(row=row, column=3).number_format = NUM_FMT
        ws.cell(row=row, column=4, value=r['not_returned']).font = DATA_FONT
        ws.cell(row=row, column=4).number_format = NUM_FMT
        ws[f'E{row}'] = f'=C{row}/B{row}'
        ws.cell(row=row, column=5).font = DATA_FONT
        ws.cell(row=row, column=5).number_format = PCT_FMT
        ws[f'F{row}'] = f'=D{row}/B{row}'
        ws.cell(row=row, column=6).font = DATA_FONT
        ws.cell(row=row, column=6).number_format = PCT_FMT
        if r['threshold'] == 12:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = HILITE_ROW
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER

    last_row = 6 + len(return_results)

    chart = LineChart()
    chart.title = "% of Accounts That Return After X Months of Inactivity"
    chart.y_axis.title = "% Returned"
    chart.x_axis.title = "Months of Inactivity"
    chart.style = 10
    chart.width = 20
    chart.height = 12
    cats = Reference(ws, min_col=1, min_row=7, max_row=last_row)
    data_ref = Reference(ws, min_col=5, min_row=6, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.width = 25000
    ws.add_chart(chart, f"A{last_row + 2}")

    # Gap cumulative distribution
    gap_start = last_row + 18
    ws.cell(row=gap_start, column=1, value="Purchase Gap Distribution").font = SUBTITLE_FONT
    ws.cell(row=gap_start + 1, column=1,
            value="Cumulative % of all purchase gaps by duration:").font = Font(
        name='Arial', size=10, italic=True)

    for col, h in enumerate(['Gap (Months)', 'Cumulative %'], 1):
        cell = ws.cell(row=gap_start + 2, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    for i, (m, pct) in enumerate(gap_cum):
        row = gap_start + 3 + i
        ws.cell(row=row, column=1, value=m).font = DATA_FONT
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=pct).font = DATA_FONT
        ws.cell(row=row, column=2).number_format = PCT_FMT
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = THIN_BORDER

    for col, w in [(1, 18), (2, 18), (3, 14), (4, 16), (5, 14), (6, 18)]:
        ws.column_dimensions[get_column_letter(col)].width = w


def build_sheet2(wb, seg_rates, sub_rates):
    ws = wb.create_sheet("Segment Analysis")

    ws['A1'] = "Dormant Analysis by Segment"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:H1')

    seg_thresholds = [6, 9, 12, 15, 18, 24]

    def write_table(start_row, title, data, label):
        ws.cell(row=start_row, column=1, value=title).font = SUBTITLE_FONT
        seg_headers = [label, '6 mo', '9 mo', '12 mo', '15 mo', '18 mo', '24 mo', 'Total Gaps']
        for col, h in enumerate(seg_headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=h)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        for i, row_data in enumerate(data):
            row = start_row + 2 + i
            ws.cell(row=row, column=1, value=row_data['group']).font = DATA_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            for j, t in enumerate(seg_thresholds):
                cell = ws.cell(row=row, column=2 + j)
                rate = row_data.get(f'{t}mo')
                if rate is not None:
                    cell.value = rate
                    cell.number_format = PCT_FMT
                    if rate >= 0.5:
                        cell.fill = GREEN_CELL
                    elif rate >= 0.25:
                        cell.fill = YELLOW_CELL
                    else:
                        cell.fill = RED_CELL
                else:
                    cell.value = "N/A"
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal='center')
                cell.border = THIN_BORDER
            ws.cell(row=row, column=8, value=row_data['total']).font = DATA_FONT
            ws.cell(row=row, column=8).number_format = NUM_FMT
            ws.cell(row=row, column=8).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=8).border = THIN_BORDER

        return start_row + 2 + len(data) + 2

    next_row = write_table(3, "Return Rate by Segment and Inactivity Threshold", seg_rates, "Segment")
    write_table(next_row, "Return Rate by Sub Segment", sub_rates, "Sub Segment")

    ws.column_dimensions['A'].width = 25
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Legend
    legend_row = next_row + len(sub_rates) + 4
    ws.cell(row=legend_row, column=1,
            value="Green = >50% return rate, Yellow = 25-50%, Red = <25%.").font = Font(
        name='Arial', size=10, italic=True)
    ws.merge_cells(f'A{legend_row}:H{legend_row}')


def build_sheet3(wb, monthly_results, months_range):
    ws = wb.create_sheet("Monthly Model")
    months_list = [str(m) for m in months_range]
    num_months = len(months_list)

    ws['A1'] = "Monthly Account Status Model"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:N1')

    ws['A2'] = ("Excludes $0 invoiced / 0 qty rows (no revenue = not a purchase). "
                "Reconciliation: Active = Prior Active \u2212 Newly Dormant + Reactivated + New Accounts")
    ws['A2'].font = Font(name='Arial', size=9, italic=True, color='666666')
    ws.merge_cells('A2:N2')

    ws['A3'] = "Dormant Threshold (months):"
    ws['A3'].font = Font(name='Arial', bold=True, size=11)
    ws['C3'] = 12
    ws['C3'].font = Font(name='Arial', bold=True, size=14, color='0000FF')
    ws['C3'].fill = HIGHLIGHT_FILL
    ws['C3'].alignment = Alignment(horizontal='center')
    ws['C3'].border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium'),
    )
    dv = DataValidation(type="list", formula1='"6,9,12,15,18,24"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws['C3'])
    ws['D3'] = "\u2190 Change this value to update the model"
    ws['D3'].font = Font(name='Arial', size=10, italic=True, color='666666')

    model_headers = [
        'Month', 'Prior Active', 'Newly Dormant', 'Reactivated', 'New Accounts',
        'Active Accounts', 'Dormant (Total)', 'Total Ever',
        'Active %', 'Dormant %',
        'Active YoY Chg', 'Active YoY %', 'New Accts YoY Chg', 'New Accts YoY %',
    ]
    recon_colors = {2: 'B4C6E7', 3: 'FFC7CE', 4: 'C6EFCE', 5: 'C6EFCE', 6: 'B4C6E7'}

    for col, h in enumerate(model_headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for col, color in recon_colors.items():
        ws.cell(row=5, column=col).fill = PatternFill('solid', fgColor=color)
        ws.cell(row=5, column=col).font = Font(name='Arial', bold=True, size=10)

    # Reference data uses 6 columns per threshold now
    for i in range(num_months):
        data_row = 6 + i
        ref_row = i + 2

        ws.cell(row=data_row, column=1, value=months_list[i]).font = DATA_FONT
        ws.cell(row=data_row, column=1).alignment = Alignment(horizontal='center')

        if i > 0:
            ws[f'B{data_row}'] = f'=F{data_row - 1}'
        else:
            ws.cell(row=data_row, column=2, value=0)
        ws.cell(row=data_row, column=2).font = DATA_FONT
        ws.cell(row=data_row, column=2).number_format = NUM_FMT

        # C: Newly Dormant (offset +4 in 6-col blocks)
        ws[f'C{data_row}'] = (f"=INDEX('Reference Data'!$A${ref_row}:$AK${ref_row},"
                               f"1,(MATCH($C$3,{{6,9,12,15,18,24}},0)-1)*6+4)")
        ws.cell(row=data_row, column=3).font = DATA_FONT
        ws.cell(row=data_row, column=3).number_format = NUM_FMT

        # D: Reactivated (+5)
        ws[f'D{data_row}'] = (f"=INDEX('Reference Data'!$A${ref_row}:$AK${ref_row},"
                               f"1,(MATCH($C$3,{{6,9,12,15,18,24}},0)-1)*6+5)")
        ws.cell(row=data_row, column=4).font = DATA_FONT
        ws.cell(row=data_row, column=4).number_format = NUM_FMT

        # E: New Accounts (+6)
        ws[f'E{data_row}'] = (f"=INDEX('Reference Data'!$A${ref_row}:$AK${ref_row},"
                               f"1,(MATCH($C$3,{{6,9,12,15,18,24}},0)-1)*6+6)")
        ws.cell(row=data_row, column=5).font = DATA_FONT
        ws.cell(row=data_row, column=5).number_format = NUM_FMT

        # F: Active (+2)
        ws[f'F{data_row}'] = (f"=INDEX('Reference Data'!$A${ref_row}:$AK${ref_row},"
                               f"1,(MATCH($C$3,{{6,9,12,15,18,24}},0)-1)*6+2)")
        ws.cell(row=data_row, column=6).font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=data_row, column=6).number_format = NUM_FMT

        # G: Dormant (+3)
        ws[f'G{data_row}'] = (f"=INDEX('Reference Data'!$A${ref_row}:$AK${ref_row},"
                               f"1,(MATCH($C$3,{{6,9,12,15,18,24}},0)-1)*6+3)")
        ws.cell(row=data_row, column=7).font = DATA_FONT
        ws.cell(row=data_row, column=7).number_format = NUM_FMT

        # H: Ever (+7)
        ws[f'H{data_row}'] = (f"=INDEX('Reference Data'!$A${ref_row}:$AK${ref_row},"
                               f"1,(MATCH($C$3,{{6,9,12,15,18,24}},0)-1)*6+7)")
        ws.cell(row=data_row, column=8).font = DATA_FONT
        ws.cell(row=data_row, column=8).number_format = NUM_FMT

        # I: Active %
        ws[f'I{data_row}'] = f'=IF(H{data_row}>0,F{data_row}/H{data_row},0)'
        ws.cell(row=data_row, column=9).font = DATA_FONT
        ws.cell(row=data_row, column=9).number_format = PCT_FMT

        # J: Dormant %
        ws[f'J{data_row}'] = f'=IF(H{data_row}>0,G{data_row}/H{data_row},0)'
        ws.cell(row=data_row, column=10).font = DATA_FONT
        ws.cell(row=data_row, column=10).number_format = PCT_FMT

        # YoY (12 months back)
        if i >= 12:
            yoy_ref = data_row - 12
            ws[f'K{data_row}'] = f'=F{data_row}-F{yoy_ref}'
            ws.cell(row=data_row, column=11).font = DATA_FONT
            ws.cell(row=data_row, column=11).number_format = '#,##0;(#,##0);"-"'
            ws[f'L{data_row}'] = f'=IF(F{yoy_ref}>0,(F{data_row}-F{yoy_ref})/F{yoy_ref},0)'
            ws.cell(row=data_row, column=12).font = DATA_FONT
            ws.cell(row=data_row, column=12).number_format = PCT_FMT
            ws[f'M{data_row}'] = f'=E{data_row}-E{yoy_ref}'
            ws.cell(row=data_row, column=13).font = DATA_FONT
            ws.cell(row=data_row, column=13).number_format = '#,##0;(#,##0);"-"'
            ws[f'N{data_row}'] = f'=IF(E{yoy_ref}>0,(E{data_row}-E{yoy_ref})/E{yoy_ref},0)'
            ws.cell(row=data_row, column=14).font = DATA_FONT
            ws.cell(row=data_row, column=14).number_format = PCT_FMT

        for col in range(1, 15):
            ws.cell(row=data_row, column=col).border = THIN_BORDER
        if i % 2 == 0:
            for col in range(1, 15):
                ws.cell(row=data_row, column=col).fill = ALT_ROW

    last_data_row = 5 + num_months

    # Chart 1: Active vs Dormant stacked bar
    chart1 = BarChart()
    chart1.type = "col"
    chart1.grouping = "stacked"
    chart1.title = "Active vs Dormant Accounts Over Time"
    chart1.y_axis.title = "Number of Accounts"
    chart1.style = 10
    chart1.width = 28
    chart1.height = 14
    cats = Reference(ws, min_col=1, min_row=6, max_row=last_data_row)
    chart1.add_data(Reference(ws, min_col=6, min_row=5, max_row=last_data_row), titles_from_data=True)
    chart1.add_data(Reference(ws, min_col=7, min_row=5, max_row=last_data_row), titles_from_data=True)
    chart1.set_categories(cats)
    chart1.series[0].graphicalProperties.solidFill = "2F5496"
    chart1.series[1].graphicalProperties.solidFill = "C00000"
    ws.add_chart(chart1, f"A{last_data_row + 3}")

    # Chart 2: Monthly flows
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "clustered"
    chart2.title = "Monthly Account Flows: New, Reactivated, and Newly Dormant"
    chart2.y_axis.title = "Number of Accounts"
    chart2.style = 10
    chart2.width = 28
    chart2.height = 14
    chart2.add_data(Reference(ws, min_col=5, min_row=5, max_row=last_data_row), titles_from_data=True)
    chart2.add_data(Reference(ws, min_col=4, min_row=5, max_row=last_data_row), titles_from_data=True)
    chart2.add_data(Reference(ws, min_col=3, min_row=5, max_row=last_data_row), titles_from_data=True)
    chart2.set_categories(cats)
    chart2.series[0].graphicalProperties.solidFill = "548235"
    chart2.series[1].graphicalProperties.solidFill = "2E75B6"
    chart2.series[2].graphicalProperties.solidFill = "C00000"
    ws.add_chart(chart2, f"A{last_data_row + 20}")

    col_widths = {1: 12, 2: 14, 3: 14, 4: 14, 5: 14, 6: 16, 7: 14,
                  8: 14, 9: 10, 10: 10, 11: 14, 12: 12, 13: 16, 14: 14}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A6'


def build_reference_sheet(wb, monthly_results, months_range):
    ws = wb.create_sheet("Reference Data")
    months_list = [str(m) for m in months_range]

    ws.cell(row=1, column=1, value="Month")
    col_idx = 2
    threshold_col_map = {}
    for t in THRESHOLDS:
        threshold_col_map[t] = col_idx
        ws.cell(row=1, column=col_idx, value=f"{t}mo_Active")
        ws.cell(row=1, column=col_idx + 1, value=f"{t}mo_Dormant")
        ws.cell(row=1, column=col_idx + 2, value=f"{t}mo_NewlyDormant")
        ws.cell(row=1, column=col_idx + 3, value=f"{t}mo_Reactivated")
        ws.cell(row=1, column=col_idx + 4, value=f"{t}mo_New")
        ws.cell(row=1, column=col_idx + 5, value=f"{t}mo_Ever")
        col_idx += 6

    for i, month in enumerate(months_list):
        row = i + 2
        ws.cell(row=row, column=1, value=month)
        for t in THRESHOLDS:
            data = monthly_results[t][i]
            base_col = threshold_col_map[t]
            ws.cell(row=row, column=base_col, value=int(data['active']))
            ws.cell(row=row, column=base_col + 1, value=int(data['dormant_total']))
            ws.cell(row=row, column=base_col + 2, value=int(data['newly_dormant']))
            ws.cell(row=row, column=base_col + 3, value=int(data['reactivated']))
            ws.cell(row=row, column=base_col + 4, value=int(data['new_accounts']))
            ws.cell(row=row, column=base_col + 5, value=int(data['ever_purchased']))


def main():
    parser = argparse.ArgumentParser(description='Dormant Account Analysis')
    parser.add_argument('account_csv', help='Path to AccountMasterData CSV')
    parser.add_argument('transaction_csv', help='Path to TransactionMasterData CSV')
    parser.add_argument('output_xlsx', help='Path for output Excel workbook')
    parser.add_argument('--start-year', type=int, default=None,
                        help='Start year for monthly model (default: auto-detect)')
    args = parser.parse_args()

    print("Loading and cleaning data...")
    acct, txn = load_and_clean(args.account_csv, args.transaction_csv)
    print(f"  Accounts: {acct['Customer Number'].nunique():,}")
    print(f"  Transactions: {len(txn):,}")

    max_period = txn['order_month'].max()
    today = max_period
    print(f"  Date range: {txn['order_month'].min()} to {max_period}")

    # Auto-detect start year: first full January
    if args.start_year is None:
        min_year = txn['Year'].min()
        if txn[txn['Year'] == min_year]['MonthNo'].min() > 1:
            start_year = min_year + 1
        else:
            start_year = min_year
    else:
        start_year = args.start_year
    print(f"  Monthly model starts: January {start_year}")

    print("\nAnalysis 1: Return rate by inactivity threshold...")
    return_results, gap_cum = analysis1_return_rates(txn, today)
    for r in return_results:
        total = r['total']
        if total > 0:
            pct = r['not_returned'] / total * 100
        else:
            pct = 0
        print(f"  {r['threshold']:>2}mo: {pct:.1f}% did not return (n={total:,})")

    print("\nAnalysis 1b: Segment breakdown...")
    seg_rates, sub_rates = analysis1_segment_rates(txn, today)
    print(f"  {len(seg_rates)} segments, {len(sub_rates)} sub-segments with sufficient data")

    print("\nAnalysis 2: Monthly model...")
    monthly_results, months_range = analysis2_monthly_model(txn, start_year)

    print("\nBuilding Excel workbook...")
    wb = Workbook()
    build_sheet1(wb, return_results, gap_cum)
    build_sheet2(wb, seg_rates, sub_rates)
    build_sheet3(wb, monthly_results, months_range)
    build_reference_sheet(wb, monthly_results, months_range)

    wb.save(args.output_xlsx)
    print(f"\nSaved: {args.output_xlsx}")
    print("NOTE: Run recalc.py on the output to calculate formula values.")


if __name__ == '__main__':
    main()
